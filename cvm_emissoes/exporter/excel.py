"""
Exportação dos dados coletados para Excel (.xlsx).

Gera duas abas:
  - "Emissões": base consolidada com todas as séries coletadas
  - "Erros": log de emissões onde algum campo não foi encontrado
"""

import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Mapeamento: chave interna -> nome da coluna no Excel
COLUNAS_EMISSOES = {
    "status":              "Status",
    "data_requerimento":   "Data de Requerimento",
    "data_encerramento":   "Data de Encerramento",
    "emissora_devedora":   "Emissora / Devedora",
    "setor":               "Setor",
    "data_emissao":        "Data de Emissão",
    "valor_mobiliario":    "Valor Mobiliário",
    "incentivada":         "Incentivada",
    "nome":                "Nome",
    "publico_alvo":        "Público-Alvo",
    "volume_serie":        "Volume da Série (R$)",
    "volume_inicial_oferta": "Volume Inicial da Oferta (R$)",
    "volume_final_oferta": "Volume Final da Oferta (R$)",
    "prazo":               "Prazo",
    "amortizacao":         "Amortização",
    "taxa_teto":           "Taxa Teto",
    "taxa_final":          "Taxa Final",
    "agencia_rating":      "Agência Avaliadora de Rating",
    "rating":              "Rating",
    "firme_sindicato":     "Firme do Sindicato",
    "book_mercado":        "Book Mercado (Qtd. VM)",
    "book_consorcio":      "Book Consórcio (Qtd. VM)",
    "coordenadores":       "Coordenadores",
    "link_cvm":            "Link CVM",
}

COLUNAS_ERROS = {
    "id_requerimento":  "ID Requerimento",
    "numero_processo":  "Número do Processo",
    "nome_emissor":     "Nome do Emissor",
    "valor_mobiliario": "Valor Mobiliário",
    "data":             "Data Requerimento",
    "erros":            "Campos com Erro",
    "link_cvm":         "Link CVM",
}

# Colunas que contêm valores monetários (para tentar converter para número)
COLUNAS_MONETARIAS = {
    "volume_serie",
    "volume_inicial_oferta",
    "volume_final_oferta",
}

# Colunas de data
COLUNAS_DATA = {
    "data_requerimento",
    "data_encerramento",
    "data_emissao",
}

# Estilo header
COR_HEADER = "1F4E79"   # azul escuro CVM
COR_FONT_HEADER = "FFFFFF"


def _parse_valor_monetario(valor_str: Optional[str]) -> Optional[float]:
    """Converte strings no formato '200.000.000,0000' para float."""
    if not valor_str:
        return None
    try:
        limpo = re.sub(r"[^\d,]", "", str(valor_str)).replace(",", ".")
        return float(limpo)
    except (ValueError, AttributeError):
        return None


def _parse_data(data_str: Optional[str]) -> Optional[datetime]:
    """Converte string DD/MM/AAAA para datetime."""
    if not data_str:
        return None
    try:
        return datetime.strptime(data_str.strip(), "%d/%m/%Y")
    except (ValueError, AttributeError):
        return None


def _escrever_header(ws, colunas: dict, cor_hex: str = COR_HEADER):
    """Escreve linha de cabeçalho com estilo."""
    fill = PatternFill("solid", fgColor=cor_hex)
    font = Font(bold=True, color=COR_FONT_HEADER)
    for col_idx, label in enumerate(colunas.values(), start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _ajustar_largura_colunas(ws, colunas: dict, min_w: int = 12, max_w: int = 60):
    """Ajusta a largura de cada coluna baseado no conteúdo."""
    for col_idx, label in enumerate(colunas.values(), start=1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(label))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
        ajustado = max(min_w, min(max_w, max_length + 2))
        ws.column_dimensions[col_letter].width = ajustado


def _escrever_linha_emissao(ws, row_idx: int, registro: dict):
    """Escreve uma linha de emissão na aba principal."""
    for col_idx, chave in enumerate(COLUNAS_EMISSOES.keys(), start=1):
        valor_raw = registro.get(chave)
        cell = ws.cell(row=row_idx, column=col_idx)

        if chave in COLUNAS_MONETARIAS:
            valor_numerico = _parse_valor_monetario(valor_raw)
            if valor_numerico is not None:
                cell.value = valor_numerico
                cell.number_format = '#,##0.00'
            else:
                cell.value = valor_raw

        elif chave in COLUNAS_DATA:
            data = _parse_data(valor_raw)
            if data:
                cell.value = data
                cell.number_format = "DD/MM/YYYY"
            else:
                cell.value = valor_raw

        elif chave == "link_cvm" and valor_raw:
            cell.value = valor_raw
            cell.hyperlink = valor_raw
            cell.style = "Hyperlink"

        elif chave in {"book_mercado", "book_consorcio"} and isinstance(valor_raw, (int, float)):
            cell.value = valor_raw
            cell.number_format = "#,##0.00##"

        else:
            cell.value = valor_raw if valor_raw is not None else ""

        cell.alignment = Alignment(vertical="top", wrap_text=(chave in {
            "coordenadores", "amortizacao", "taxa_teto", "taxa_final",
            "emissora_devedora",
        }))


def exportar_excel(
    registros: list[dict],
    erros: list[dict],
    output_path: Optional[str] = None,
) -> str:
    """
    Exporta os registros coletados para um arquivo Excel.

    Args:
        registros: Lista de dicionários com dados das séries
        erros: Lista de dicionários com log de erros
        output_path: Caminho do arquivo de saída (opcional, usa padrão com timestamp)

    Returns:
        Caminho absoluto do arquivo gerado
    """
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(Path.cwd() / f"emissoes_cvm_{ts}.xlsx")

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------
    # Aba principal: Emissões
    # ------------------------------------------------------------------
    ws_emissoes = wb.active
    ws_emissoes.title = "Emissões"
    ws_emissoes.freeze_panes = "A2"  # congela cabeçalho

    _escrever_header(ws_emissoes, COLUNAS_EMISSOES)

    for row_idx, registro in enumerate(registros, start=2):
        _escrever_linha_emissao(ws_emissoes, row_idx, registro)

    # Altura padrão para linhas de dados
    for row in ws_emissoes.iter_rows(min_row=2):
        ws_emissoes.row_dimensions[row[0].row].height = 30

    # Altura do cabeçalho
    ws_emissoes.row_dimensions[1].height = 35

    _ajustar_largura_colunas(ws_emissoes, COLUNAS_EMISSOES)

    # ------------------------------------------------------------------
    # Aba de erros
    # ------------------------------------------------------------------
    if erros:
        ws_erros = wb.create_sheet("Erros")
        _escrever_header(ws_erros, COLUNAS_ERROS, cor_hex="8B0000")  # vermelho escuro

        for row_idx, erro in enumerate(erros, start=2):
            for col_idx, chave in enumerate(COLUNAS_ERROS.keys(), start=1):
                valor = erro.get(chave, "")
                cell = ws_erros.cell(row=row_idx, column=col_idx, value=valor)
                if chave == "link_cvm" and valor:
                    cell.hyperlink = valor
                    cell.style = "Hyperlink"
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        _ajustar_largura_colunas(ws_erros, COLUNAS_ERROS)
        ws_erros.row_dimensions[1].height = 35

    # ------------------------------------------------------------------
    # Salvar
    # ------------------------------------------------------------------
    wb.save(output_path)
    return str(output_path)
